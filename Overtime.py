from __future__ import division
from collections import deque
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import sys
import requests
import json
import re
import datetime

excel_filename = 'rc_pm.xlsx'

harvest_headers = {
    'Content-type': 'application/json',
    'Accept': 'application/json',
    'Authorization': ''
}

contractors = ['Alfonso', 'Amos', 'Luke', 'Rex', 'Richard']

def init():
    # Harvest - Build request & load projects
    projects = requests.get('https://revacomm.harvestapp.com/projects', headers=harvest_headers)
    projects_json = projects.json()

    return projects_json


def openExcel(filename):
    wb = None
    ws = None

    # Open or create a new worksheet
    today = datetime.datetime.now().strftime("%Y.%m.%d")
    if os.path.exists(filename):
        wb = load_workbook(filename)
        if today in wb.get_sheet_names():
            ws = wb[today]
    else:
        wb = Workbook()
        ws = wb.active

    # Delete current sheet & create new
    if ws is not None:
        wb.remove_sheet(ws)
    ws = wb.create_sheet(0)
    ws.title = today

    # CM setup sheet headers
    headers = ["Harvest Code", "Wrike Name", "Completion", "Burn", "Remain"]
    for col, header in enumerate(headers):
        c = ws.cell(row = 1, column = col+1)
        c.value = header
    return (wb, ws)


def closeExcel(wb, filename):
    wb.save(filename)


def outputToExcel(ws, project, index):
    proj_tmp = [project["Harvest_Code"], project["Wrike_Name"], project["Progress"]["Completion"], project["Progress"]["Burn"], project["Progress"]["Remain"]]
    for col, val in enumerate(proj_tmp):
        c = ws.cell(row = index, column = col+1)
        c.value = val


def userTotalTime(userTimeJson, isContract):
    hours = 0
    timeByDay = {}
    for timeEntry in userTimeJson:
        timeEntry = timeEntry["day_entry"]
        spentAt = timeEntry["spent_at"]

        isWeekend = False
        date = datetime.datetime.strptime(spentAt, '%Y-%m-%d')
        if not isContract:
            if date.weekday() >= 5:
                isWeekend = True
        else:  # Contractors work our Tue-Sat
            if date.weekday() == 0 or date.weekday() == 6:
                isWeekend = True

        if spentAt not in timeByDay:
            timeByDay[spentAt] = {
                "weekend": isWeekend,
                "hours": 0
            }
        timeByDay[spentAt]["hours"] += timeEntry["hours"]
        hours = hours + timeEntry["hours"]

    over = 0
    under = 0
    for date in timeByDay:
        weekend = timeByDay[date]["weekend"]
        dailyHours = timeByDay[date]["hours"]
        if weekend:
            over += dailyHours
        elif dailyHours < 8:
            under += (8 - dailyHours)
        elif dailyHours > 8:
            over += (dailyHours - 8)
    over -= under
    if over < 0:
        over = 0
    return (hours, over)


if __name__ == '__main__':
    firstDayOfYear = '20160101'
    today = str(datetime.datetime.today().strftime('%Y%m%d'))
    
    contractOver = 0
    coreOver = 0
    peopleTime = {}
    people = requests.get('https://revacomm.harvestapp.com/people', headers=harvest_headers)
    people_json = people.json()
    for person in people_json:
        pUser = person['user']
        uid = str(pUser['id'])
        first = pUser['first_name']
        last = pUser['last_name']

        userTime = requests.get('https://revacomm.harvestapp.com/people/' + uid + '/entries?from=' + firstDayOfYear + '&to=' + today, headers=harvest_headers)
        userTime_json = userTime.json()
        if not userTime_json:
            continue
        isContract = first in contractors
        hours, over = userTotalTime(userTime_json, isContract)
        if isContract:
            contractOver += over
        else:
            coreOver += over
        print first + ": TOT: " + str(hours) + " OVER: " + str(over)

        peopleTime[uid] = {
            "first": first,
            "last": last,
            "total_hours": hours,
            "overtime": over
        }

    print "Core Over: " + str(coreOver)
    print "Contract Over: " + str(contractOver)
    print "Total Over: " + str(coreOver + contractOver)
